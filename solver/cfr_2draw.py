#!/usr/bin/env python3
# this is the 2-draw model code

from __future__ import annotations
import argparse
import os
import random
import numpy as np
from dataclasses import dataclass
from typing import Dict, Callable, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from cfr_common_bucketgame import (
    SeedItem,
    reset_infosets,
    build_unique_seeds_and_weights_from_list,
    build_bucket_pair_matchups_seedmodel,
    BUCKETS_2DRAW,
    bucket_label_2draw,
    train_cfrplus_sweep,
    export_variant as common_export_variant,
    order_actions_limit,
    A_CHECK, A_BET, A_CALL, A_FOLD, A_RAISE,
    Matchup,
    infosets,
    eval_chip_btn,
)

IGNORED_EV_BUCKETS = {"2-Pair", "Trips", "Extra"}

# labels used when exporting action names
ACTION_LABELS = {
    A_CHECK: "Check",
    A_BET: "Bet",
    A_CALL: "Call",
    A_FOLD: "Fold",
    A_RAISE: "Raise",
}


class _PrefixSampler:
    def __init__(self, items: List[Matchup], weights: List[float]):
        if len(items) != len(weights):
            raise ValueError("items and weights length mismatch")
        self.items = list(items)
        self.prefix = []
        s = 0.0
        for w in weights:
            s += max(0.0, float(w))
            self.prefix.append(s)
        self.total = s

    def sample(self) -> Matchup:
        if not self.items or self.total <= 0:
            raise RuntimeError("PrefixSampler has no mass/items")
        x = random.random() * self.total
        import bisect
        idx = bisect.bisect_left(self.prefix, x)
        if idx < 0:
            idx = 0
        if idx >= len(self.items):
            idx = len(self.items) - 1
        return self.items[idx]


def _build_conditional_matchup_samplers(
    matchups: List[Matchup],
    buckets: List[str],
    matchup_weights: Optional[List[float]] = None,
) -> Dict[Tuple[int, str], Optional[_PrefixSampler]]:
    by_key: Dict[Tuple[int, str], List[Matchup]] = {}
    w_key: Dict[Tuple[int, str], List[float]] = {}

    allowed = set(buckets)

    for b in buckets:
        by_key[(0, b)] = []
        w_key[(0, b)] = []
        by_key[(1, b)] = []
        w_key[(1, b)] = []

    if matchup_weights is not None and len(matchup_weights) != len(matchups):
        raise ValueError("matchup_weights must be None or the same length as matchups")

    for i, m in enumerate(matchups):
        w = float(matchup_weights[i]) if matchup_weights is not None else float(m.prob)
        if w <= 0.0:
            continue

        if m.btn_bucket in allowed:
            by_key[(0, m.btn_bucket)].append(m)
            w_key[(0, m.btn_bucket)].append(w)

        if m.bb_bucket in allowed:
            by_key[(1, m.bb_bucket)].append(m)
            w_key[(1, m.bb_bucket)].append(w)

    out: Dict[Tuple[int, str], Optional[_PrefixSampler]] = {}
    for k in by_key.keys():
        items = by_key[k]
        ws = w_key[k]
        if not items or sum(ws) <= 0.0:
            out[k] = None
            continue
        out[k] = _PrefixSampler(items, ws)
    return out


def _safe_sheet_title(title: str, used: set) -> str:
    bad = set(r'[]:*?/\ ')
    cleaned = ''.join('_' if ch in bad else ch for ch in title)
    cleaned = cleaned if cleaned else 'root'
    cleaned = cleaned[:31]

    base = cleaned
    n = 1
    while cleaned in used:
        suffix = f'_{n}'
        cleaned = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(cleaned)
    return cleaned


def _style_ws_header(ws, row_idx: int, ncols: int) -> None:
    fill = PatternFill('solid', fgColor='1F4E78')
    font = Font(color='FFFFFF', bold=True)
    align = Alignment(horizontal='center', vertical='center')
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _autosize_ws(ws, min_w: int = 10, max_w: int = 18) -> None:
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = '' if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, min_w), len(val) + 2)

    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_w, min(max_w, w))


def compute_action_ev_rows_for_state(
    s0,
    *,
    buckets: List[str],
    bucket_freq: Dict[str, float],
    matchups: List[Matchup],
    samples_per_bucket: int,
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    matchup_weights: Optional[List[float]] = None,
) -> dict:
    root_pot = float(s0.pot)
    root_inv_btn = float(s0.invested_btn)
    root_inv_bb = float(s0.invested_bb)

    actor = int(s0.to_act)
    actor_name = 'BTN' if actor == 0 else 'BB'
    acts = legal_actions_fn(s0)

    samplers = _build_conditional_matchup_samplers(matchups, buckets, matchup_weights=matchup_weights)

    rows = []
    total_rate = 0.0
    overall_node_btn_num = 0.0
    overall_action_btn_num = {a: 0.0 for a in acts}

    for b in buckets:
        rate = float(bucket_freq.get(b, 0.0))
        rate_pct = 100.0 * rate
        samp = samplers.get((actor, b))

        row = {'bucket': b, 'rate': round(rate_pct, 6), 'btn_ev': None, 'bb_ev': None}
        for a in acts:
            row[f'{a}_btn_ev'] = None
            row[f'{a}_bb_ev'] = None

        if samp is None or rate <= 0.0:
            rows.append(row)
            continue

        total_node_btn = 0.0
        action_btn_totals = {a: 0.0 for a in acts}

        for _ in range(int(samples_per_bucket)):
            m = samp.sample()
            total_node_btn += eval_chip_btn(
                m, s0, root_pot, root_inv_btn, root_inv_bb,
                legal_actions_fn, step_fn, ikey_fn
            )
            for a in acts:
                ns = step_fn(s0, a)
                action_btn_totals[a] += eval_chip_btn(
                    m, ns, root_pot, root_inv_btn, root_inv_bb,
                    legal_actions_fn, step_fn, ikey_fn
                )

        btn_ev = total_node_btn / float(samples_per_bucket)
        bb_ev = float(root_pot) - float(btn_ev)

        row['btn_ev'] = round(float(btn_ev), 6)
        row['bb_ev'] = round(float(bb_ev), 6)

        total_rate += rate
        overall_node_btn_num += rate * btn_ev

        for a in acts:
            a_btn = action_btn_totals[a] / float(samples_per_bucket)
            a_bb = float(root_pot) - float(a_btn)
            row[f'{a}_btn_ev'] = round(float(a_btn), 6)
            row[f'{a}_bb_ev'] = round(float(a_bb), 6)
            overall_action_btn_num[a] += rate * a_btn

        rows.append(row)

    overall = {
        'actor': actor_name,
        'history': getattr(s0, 'history', ''),
        'pot': float(s0.pot),
        'to_call': float(getattr(s0, 'to_call', 0.0)),
        'raises_made': int(getattr(s0, 'raises_made', 0)),
        'reach_freq_pct': round(100.0 * total_rate, 6),
        'btn_ev': None,
        'bb_ev': None,
        'actions': acts[:],
    }

    if total_rate > 0.0:
        overall_btn = overall_node_btn_num / total_rate
        overall['btn_ev'] = round(float(overall_btn), 6)
        overall['bb_ev'] = round(float(root_pot - overall_btn), 6)
        for a in acts:
            a_btn = overall_action_btn_num[a] / total_rate
            overall[f'{a}_btn_ev'] = round(float(a_btn), 6)
            overall[f'{a}_bb_ev'] = round(float(root_pot - a_btn), 6)
    else:
        for a in acts:
            overall[f'{a}_btn_ev'] = None
            overall[f'{a}_bb_ev'] = None

    return {'actor': actor_name, 'actions': acts[:], 'rows': rows, 'overall': overall}


def compute_bucket_freq_by_sequence(
    *,
    matchups: List[Matchup],
    buckets: List[str],
    report_seqs: List[str],
    pot: float,
    bucket_freq_by_player: Dict[str, Dict[str, float]],
    initial_state_fn: Callable[[float], object],
    state_from_history_fn: Callable[[str, float], object],
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
) -> Tuple[Dict[str, dict], Dict[str, List[float]]]:
    report_set = set(report_seqs)
    seq_info: Dict[str, dict] = {}
    reach_prob: Dict[str, List[float]] = {seq: [0.0] * len(matchups) for seq in report_seqs}

    allowed = set(buckets)

    for seq in report_seqs:
        s0 = state_from_history_fn(seq, float(pot))
        if getattr(s0, 'terminal', False):
            seq_info[seq] = {'terminal': True}
            continue
        actor = 'BTN' if int(s0.to_act) == 0 else 'BB'
        seq_info[seq] = {
            'actor': actor,
            'bucket_freq': {b: 0.0 for b in buckets},
            'bucket_reach': {b: 0.0 for b in buckets},
            'total_freq': 0.0,
        }

    for i, m in enumerate(matchups):
        root = initial_state_fn(float(pot))

        def walk(s, pr: float) -> None:
            if getattr(s, 'terminal', False):
                return

            h = getattr(s, 'history', '')
            if h in report_set:
                info = seq_info.get(h)
                if info is not None and not info.get('terminal', False):
                    reach_prob[h][i] += float(pr)
                    actor = info['actor']
                    b = m.btn_bucket if actor == 'BTN' else m.bb_bucket
                    if b in allowed:
                        info['bucket_freq'][b] = float(info['bucket_freq'].get(b, 0.0)) + float(m.prob) * float(pr)
                        info['total_freq'] = float(info['total_freq']) + float(m.prob) * float(pr)

            p = int(s.to_act)
            bucket = m.btn_bucket if p == 0 else m.bb_bucket
            acts = legal_actions_fn(s)
            node = infosets.get(ikey_fn(p, bucket, s))
            if node is None:
                probs = [1.0 / max(1, len(acts))] * len(acts)
            else:
                avg = node.avg_strategy()
                probs = list(avg) if len(avg) == len(acts) else [1.0 / max(1, len(acts))] * len(acts)

            for a, pa in zip(acts, probs):
                pa = float(pa)
                if pa <= 0.0:
                    continue
                walk(step_fn(s, a), float(pr) * pa)

        walk(root, 1.0)

    for seq, info in seq_info.items():
        if info.get('terminal', False):
            continue
        actor = info['actor']
        marg = bucket_freq_by_player.get(actor, {})
        for b in buckets:
            denom = float(marg.get(b, 0.0))
            num = float(info['bucket_freq'].get(b, 0.0))
            info['bucket_reach'][b] = (num / denom) if denom > 0 else 0.0

    return seq_info, reach_prob


def export_move_evs_excel(
    *,
    out_prefix: str,
    pot: float,
    report_seqs: List[str],
    buckets: List[str],
    matchups: List[Matchup],
    seq_info: Dict[str, dict],
    reach_prob: Dict[str, List[float]],
    state_from_history_fn: Callable[[str, float], object],
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    order_actions_fn: Callable[[List[str]], List[str]],
    ev_samples_per_bucket: int,
) -> None:
    out_dir = os.path.join('..', 'data', str(out_prefix), 'excel')
    os.makedirs(out_dir, exist_ok=True)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    used_titles = {'Summary'}
    summary_rows = []
    all_actions = [A_CHECK, A_BET, A_CALL, A_FOLD, A_RAISE]

    for seq in report_seqs:
        s0 = state_from_history_fn(seq, float(pot))
        seq_label = seq if seq else 'root'

        if getattr(s0, 'terminal', False):
            summary_rows.append({
                'sequence': seq_label,
                'actor': 'terminal',
                'reach_freq_pct': None,
                'btn_ev': None,
                'bb_ev': None
            })
            continue

        pr_list = reach_prob.get(seq, [0.0] * len(matchups))
        w_hist = [float(m.prob) * float(pr_list[i]) for i, m in enumerate(matchups)]
        freq_map = seq_info.get(seq, {}).get('bucket_freq', {})

        pack = compute_action_ev_rows_for_state(
            s0,
            buckets=buckets,
            bucket_freq=freq_map,
            matchups=matchups,
            samples_per_bucket=int(ev_samples_per_bucket),
            legal_actions_fn=legal_actions_fn,
            step_fn=step_fn,
            ikey_fn=ikey_fn,
            matchup_weights=w_hist,
        )

        ordered_actions = order_actions_fn(pack['actions'])
        sheet_name = _safe_sheet_title(f'seq_{seq_label}', used_titles)
        ws = wb.create_sheet(title=sheet_name)

        ws['A1'] = 'Sequence'
        ws['B1'] = seq_label
        ws['A2'] = 'Actor'
        ws['B2'] = pack['overall']['actor']
        ws['A3'] = 'Pot'
        ws['B3'] = pack['overall']['pot']
        ws['D1'] = 'To Call'
        ws['E1'] = pack['overall']['to_call']
        ws['D2'] = 'Raises Made'
        ws['E2'] = pack['overall']['raises_made']
        ws['D3'] = 'Reach Freq %'
        ws['E3'] = pack['overall']['reach_freq_pct']

        header = ['Bucket', 'Rate %', 'BTN EV', 'BB EV']
        for a in ordered_actions:
            label = ACTION_LABELS.get(a, a)
            header.extend([f'{label} BTN EV', f'{label} BB EV'])

        start_row = 5
        for j, h in enumerate(header, start=1):
            ws.cell(row=start_row, column=j, value=h)
        _style_ws_header(ws, start_row, len(header))

        r = start_row + 1
        for row in pack['rows']:
            out_vals = [row['bucket'], row['rate'], row['btn_ev'], row['bb_ev']]
            for a in ordered_actions:
                out_vals.extend([row[f'{a}_btn_ev'], row[f'{a}_bb_ev']])
            for c, v in enumerate(out_vals, start=1):
                ws.cell(row=r, column=c, value=v)
            r += 1

        r += 1
        ws.cell(row=r, column=1, value='OVERALL')
        ws.cell(row=r, column=2, value=pack['overall']['reach_freq_pct'])
        ws.cell(row=r, column=3, value=pack['overall']['btn_ev'])
        ws.cell(row=r, column=4, value=pack['overall']['bb_ev'])

        c = 5
        for a in ordered_actions:
            ws.cell(row=r, column=c, value=pack['overall'][f'{a}_btn_ev'])
            ws.cell(row=r, column=c + 1, value=pack['overall'][f'{a}_bb_ev'])
            c += 2

        for c in range(1, len(header) + 1):
            ws.cell(row=r, column=c).font = Font(bold=True)

        _autosize_ws(ws)

        sm = {
            'sequence': seq_label,
            'actor': pack['overall']['actor'],
            'reach_freq_pct': pack['overall']['reach_freq_pct'],
            'btn_ev': pack['overall']['btn_ev'],
            'bb_ev': pack['overall']['bb_ev'],
        }
        for a in all_actions:
            sm[f'{a}_btn_ev'] = pack['overall'].get(f'{a}_btn_ev')
            sm[f'{a}_bb_ev'] = pack['overall'].get(f'{a}_bb_ev')
        summary_rows.append(sm)

    summary_header = ['Sequence', 'Actor', 'Reach Freq %', 'BTN EV', 'BB EV']
    for a in all_actions:
        label = ACTION_LABELS.get(a, a)
        summary_header.extend([f'{label} BTN EV', f'{label} BB EV'])

    for j, h in enumerate(summary_header, start=1):
        ws_summary.cell(row=1, column=j, value=h)
    _style_ws_header(ws_summary, 1, len(summary_header))

    rr = 2
    for row in summary_rows:
        vals = [
            row.get('sequence'),
            row.get('actor'),
            row.get('reach_freq_pct'),
            row.get('btn_ev'),
            row.get('bb_ev')
        ]
        for a in all_actions:
            vals.extend([row.get(f'{a}_btn_ev'), row.get(f'{a}_bb_ev')])
        for c, v in enumerate(vals, start=1):
            ws_summary.cell(row=rr, column=c, value=v)
        rr += 1

    _autosize_ws(ws_summary, max_w=20)
    out_file = os.path.join(out_dir, f'move_evs_pot{int(pot)}.xlsx')
    wb.save(out_file)
    print(f'excel export -> {out_file}')


def export_variant(
    *,
    prefix: str,
    pot: float,
    iterations: int,
    buckets: List[str],
    report_seqs: List[str],
    bucket_freq_by_player: Dict[str, Dict[str, float]],
    tracked_iters: List[int],
    regret_data: dict,
    evo_data: dict,
    node_actions: dict,
    exploit_iters: List[int],
    expl_mbb: List[float],
    expl_bb: List[float],
    expl_chip: List[float],
    initial_state_fn: Callable[[float], object],
    state_from_history_fn: Callable[[str, float], object],
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    order_actions_fn: Callable[[List[str]], List[str]],
    matchups: List[Matchup],
    ev_samples_per_bucket: int,
    time_export: bool = False,
) -> None:
    common_export_variant(
        prefix=prefix,
        pot=pot,
        iterations=iterations,
        buckets=buckets,
        report_seqs=report_seqs,
        bucket_freq_by_player=bucket_freq_by_player,
        tracked_iters=tracked_iters,
        regret_data=regret_data,
        evo_data=evo_data,
        node_actions=node_actions,
        exploit_iters=exploit_iters,
        expl_mbb=expl_mbb,
        expl_bb=expl_bb,
        expl_chip=expl_chip,
        initial_state_fn=initial_state_fn,
        state_from_history_fn=state_from_history_fn,
        legal_actions_fn=legal_actions_fn,
        step_fn=step_fn,
        ikey_fn=ikey_fn,
        order_actions_fn=order_actions_fn,
        matchups=matchups,
        ev_samples_per_bucket=ev_samples_per_bucket,
        time_export=time_export,
    )

    excel_buckets = [b for b in buckets if b not in IGNORED_EV_BUCKETS]

    seq_info, reach_prob = compute_bucket_freq_by_sequence(
        matchups=matchups,
        buckets=excel_buckets,
        report_seqs=report_seqs,
        pot=float(pot),
        bucket_freq_by_player=bucket_freq_by_player,
        initial_state_fn=initial_state_fn,
        state_from_history_fn=state_from_history_fn,
        legal_actions_fn=legal_actions_fn,
        step_fn=step_fn,
        ikey_fn=ikey_fn,
    )

    export_move_evs_excel(
        out_prefix=prefix,
        pot=float(pot),
        report_seqs=report_seqs,
        buckets=excel_buckets,
        matchups=matchups,
        seq_info=seq_info,
        reach_prob=reach_prob,
        state_from_history_fn=state_from_history_fn,
        legal_actions_fn=legal_actions_fn,
        step_fn=step_fn,
        ikey_fn=ikey_fn,
        order_actions_fn=order_actions_fn,
        ev_samples_per_bucket=int(ev_samples_per_bucket),
    )


# config
SEED = 42
ITERATIONS = 50_000
POT_SIZES = [7.0]

STARTING_PLAYER = 1  # BB acts first

MAX_RAISES = 3
RAISE_SCHEDULE = [1.0, 2.0, 3.0, 4.0]  # absolute invest levels (limit-style)
REPORT_SEQS = ["", "k", "b", "kb", "kbr", "br", "brr", "brrr", "kbrr", "kbrrr"]

BIG_BLIND = 0.5
BR_POLICY_ITERS = 16
EXP_PRINT_STEP = 100
EV_SAMPLES_PER_BUCKET = 800

# 3-card seed ranges (duplicates => weights)
BTN_SEEDS_RAW = [
    "432", "532", "542", "543", "632", "642", "643", "652", "653",
    "732", "742", "743", "752", "753", "754", "762", "763",
    "832", "842", "843", "852", "853", "854", "862", "863", "864", "865",
]
BB_SEEDS_RAW = [
    "543", "643", "652", "653", "753", "754", "762", "763",
    "832", "842", "843", "852", "853", "854", "862", "863", "864", "865",
]


# limit betting game definition
@dataclass
class State:
    to_act: int
    history: str
    pot: float
    to_call: float
    raises_made: int
    terminal: bool
    winner: Optional[int]
    showdown: bool
    invested_btn: float
    invested_bb: float


def initial_state(start_pot: float) -> State:
    return State(
        to_act=STARTING_PLAYER,
        history="",
        pot=float(start_pot),
        to_call=0.0,
        raises_made=0,
        terminal=False,
        winner=None,
        showdown=False,
        invested_btn=0.0,
        invested_bb=0.0,
    )


def legal_actions(s: State) -> List[str]:
    if s.terminal:
        return []
    if s.to_call == 0:
        return [A_CHECK, A_BET]
    acts = [A_FOLD, A_CALL]
    if s.raises_made < MAX_RAISES:
        acts.append(A_RAISE)
    return acts


def _add_invested(st: State, player: int, amt: float) -> None:
    if player == 0:
        st.invested_btn += amt
    else:
        st.invested_bb += amt


def step(s: State, a: str) -> State:
    t = State(**vars(s))
    t.history += a
    p, opp = s.to_act, 1 - s.to_act

    if s.to_call == 0:
        if a == A_CHECK:
            if t.history.endswith("kk"):
                t.terminal = True
                t.showdown = True
            else:
                t.to_act = opp
        elif a == A_BET:
            amt = float(RAISE_SCHEDULE[0])
            t.pot += amt
            _add_invested(t, p, amt)
            t.to_call = amt
            t.raises_made = 0
            t.to_act = opp
    else:
        if a == A_FOLD:
            t.terminal = True
            t.winner = opp
        elif a == A_CALL:
            t.pot += s.to_call
            _add_invested(t, p, s.to_call)
            t.to_call = 0.0
            t.terminal = True
            t.showdown = True
        elif a == A_RAISE:
            newlvl = float(RAISE_SCHEDULE[s.raises_made + 1])
            curr = t.invested_btn if p == 0 else t.invested_bb
            oppi = t.invested_bb if p == 0 else t.invested_btn
            add = newlvl - curr
            t.pot += add
            _add_invested(t, p, add)
            t.to_call = newlvl - oppi
            t.raises_made = s.raises_made + 1
            t.to_act = opp

    return t


def state_from_history(h: str, start_pot: float) -> State:
    s = initial_state(start_pot)
    for ch in h:
        s = step(s, ch)
    return s


def ikey(p: int, bucket: str, s: State) -> tuple:
    return (p, bucket, s.history, s.to_call, s.raises_made)


def info_key_str(s: State) -> str:
    return f"h:{s.history}|tc:{s.to_call}|rm:{s.raises_made}"


# main
def main():
    global SEED, ITERATIONS, BR_POLICY_ITERS, EXP_PRINT_STEP, EV_SAMPLES_PER_BUCKET

    ap = argparse.ArgumentParser(description="CFR+ 2draw (FULL chance sweep bucket-game)")
    ap.add_argument("--seed", type=int, default=SEED)
    ap.add_argument("--iters", type=int, default=ITERATIONS)
    # run_all compatibility (ignored)
    ap.add_argument("--exploit-step", type=int, default=None)
    ap.add_argument("--tracking-step", type=int, default=None)
    ap.add_argument("--exploit-samples", type=int, default=None)
# real knobs
    ap.add_argument("--br-policy-iters", type=int, default=BR_POLICY_ITERS)
    ap.add_argument("--exp-print-step", type=int, default=EXP_PRINT_STEP)
    ap.add_argument("--ev-samples", type=int, default=EV_SAMPLES_PER_BUCKET)
    ap.add_argument("--time-export", action="store_true", help="Print wall-clock time for export")
    args = ap.parse_args()

    SEED = int(args.seed)
    ITERATIONS = int(args.iters)
    BR_POLICY_ITERS = int(args.br_policy_iters)
    EXP_PRINT_STEP = int(args.exp_print_step)
    EV_SAMPLES_PER_BUCKET = int(args.ev_samples)

    random.seed(SEED)
    np.random.seed(SEED)

    print(f"2draw CFR+ FULL: iters={ITERATIONS:,} seed={SEED} BB={BIG_BLIND}")

    btn_seeds, btn_w = build_unique_seeds_and_weights_from_list(BTN_SEEDS_RAW)
    bb_seeds, bb_w = build_unique_seeds_and_weights_from_list(BB_SEEDS_RAW)

    btn_items = [SeedItem(seed=s, draws=2, weight=float(w)) for s, w in zip(btn_seeds, btn_w)]
    bb_items = [SeedItem(seed=s, draws=2, weight=float(w)) for s, w in zip(bb_seeds, bb_w)]

    print("building bucket-pair chance distribution (exact draw2)...")
    matchups, bucket_freq_by_player = build_bucket_pair_matchups_seedmodel(
        btn_items, bb_items, BUCKETS_2DRAW, bucket_label_2draw
    )
    print(f"done: {len(matchups)} nonzero bucket-pairs")

    for pot in POT_SIZES:
        print(f"\npot={pot}")
        random.seed(SEED)
        np.random.seed(SEED)
        reset_infosets()

        tracked_iters, regret_data, evo_data, node_actions, ex_i, ex_mbb, ex_bb, ex_chip = train_cfrplus_sweep(
            matchups=matchups,
            buckets=BUCKETS_2DRAW,
            report_seqs=REPORT_SEQS,
            iterations=ITERATIONS,
            pot=float(pot),
            initial_state_fn=initial_state,
            state_from_history_fn=state_from_history,
            legal_actions_fn=legal_actions,
            step_fn=step,
            ikey_fn=ikey,
            info_key_str_fn=info_key_str,
            br_policy_iters=BR_POLICY_ITERS,
            big_blind=BIG_BLIND,
            exp_print_step=EXP_PRINT_STEP,
        )

        export_variant(
            prefix="2draw",
            pot=float(pot),
            iterations=ITERATIONS,
            buckets=BUCKETS_2DRAW,
            report_seqs=REPORT_SEQS,
            bucket_freq_by_player=bucket_freq_by_player,
            tracked_iters=tracked_iters,
            regret_data=regret_data,
            evo_data=evo_data,
            node_actions=node_actions,
            exploit_iters=ex_i,
            expl_mbb=ex_mbb,
            expl_bb=ex_bb,
            expl_chip=ex_chip,
            initial_state_fn=initial_state,
            state_from_history_fn=state_from_history,
            legal_actions_fn=legal_actions,
            step_fn=step,
            ikey_fn=ikey,
            order_actions_fn=order_actions_limit,
            matchups=matchups,
            ev_samples_per_bucket=EV_SAMPLES_PER_BUCKET,
            time_export=bool(args.time_export),
        )

    print("\n2draw done!")


if __name__ == "__main__":
    main()
