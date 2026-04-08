#!/usr/bin/env python3
# this is the code for the 1-draw model

from __future__ import annotations
import os
import json
import random
import bisect
import time
import argparse
from dataclasses import dataclass
from collections import Counter
from typing import Callable, Dict, List, Optional, Sequence, Tuple
import numpy as np
from math import comb

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Rank / deck helpers
RANKS: List[int] = list(range(2, 15))  # 2..14 (A)
_COPIES_PER_RANK: int = 4
_DECK_COUNTS_BASE: Tuple[int, ...] = tuple([_COPIES_PER_RANK] * len(RANKS))  # 13-long

_RANK_TO_IDX = {r: r - 2 for r in RANKS}
_IDX_TO_RANK = {i: i + 2 for i in range(len(RANKS))}


# turns one rank character into the integer version
def parse_rank_char(ch: str) -> int:
    ch = ch.strip().upper()
    if ch == "T":
        return 10
    if ch == "J":
        return 11
    if ch == "Q":
        return 12
    if ch == "K":
        return 13
    if ch == "A":
        return 14
    return int(ch)


# reads a seed string and sorts it into rank order
def parse_seed_str(s: str) -> Tuple[int, ...]:
    s = s.strip()
    return tuple(sorted(parse_rank_char(ch) for ch in s))


# counts repeated seeds in the grid so they can be weighted properly
def build_unique_seeds_and_weights_from_grid(
    grid: Sequence[Sequence[str]],
) -> Tuple[List[Tuple[int, ...]], np.ndarray]:
    flat = [h for row in grid for h in row]
    counts = Counter(parse_seed_str(h) for h in flat)
    seeds = list(counts.keys())
    weights = np.array([counts[s] for s in seeds], dtype=float)
    return seeds, weights


# removes the seed cards from the rank deck before drawing
def _counts_after_seeds(btn_seed: Tuple[int, ...], bb_seed: Tuple[int, ...]) -> Optional[Tuple[int, ...]]:
    c = list(_DECK_COUNTS_BASE)
    for r in btn_seed:
        i = _RANK_TO_IDX.get(r)
        if i is None:
            return None
        c[i] -= 1
        if c[i] < 0:
            return None
    for r in bb_seed:
        i = _RANK_TO_IDX.get(r)
        if i is None:
            return None
        c[i] -= 1
        if c[i] < 0:
            return None
    return tuple(c)


# 2-7 Lowball eval
def is_straight(vals: Tuple[int, ...]) -> bool:
    sv = sorted(set(vals))
    if len(sv) < 5:
        return False
    for i in range(len(sv) - 4):
        if sv[i + 4] - sv[i] == 4:
            return True
    return False


# classifies a rank-only 2-7 hand into its hand type
def classify_27(hand: Tuple[int, ...]) -> Tuple[int, Tuple[int, ...]]:
    c = Counter(hand)
    f = sorted(c.values(), reverse=True)
    desc = tuple(sorted(hand, reverse=True))
    if 4 in f:
        return 6, desc
    if 3 in f and 2 in f:
        return 5, desc
    if is_straight(hand) and max(f) == 1:
        return 4, desc
    if 3 in f:
        return 3, desc
    if f.count(2) == 2:
        return 2, desc
    if 2 in f:
        return 1, desc
    return 0, desc


# compares btn and bb hands and returns who wins
def compare(btn: Tuple[int, ...], bb: Tuple[int, ...]) -> int:
    c1, v1 = classify_27(btn)
    c2, v2 = classify_27(bb)
    if c1 < c2:
        return +1
    if c1 > c2:
        return -1
    if v1 < v2:
        return +1
    if v1 > v2:
        return -1
    return 0


# Buckets
_EXTRA_BUCKET = "Extra"
STRAIGHT_BUCKET = "Straight"

BUCKETS_1DRAW: List[str] = [
    "75", "76", "85", "86", "87", "95", "96", "97", "98",
    "T5", "T6", "T7", "T8", "T9", "J8", "J9", "Q", "K", "A",
    "22", "33", "44", "55", "66", "77", "88", "99",
    STRAIGHT_BUCKET,
]


# maps a high-card hand into the right bucket label
def _highcard_bucket_label(hi: int, sh: int) -> str:
    if hi == 7:
        if sh <= 5:
            return "75"
        if sh == 6:
            return "76"
        return _EXTRA_BUCKET

    if hi == 8:
        if sh == 5:
            return "85"
        if sh == 6:
            return "86"
        if sh == 7:
            return "87"
        return _EXTRA_BUCKET

    if hi == 9:
        if sh == 5:
            return "95"
        if sh == 6:
            return "96"
        if sh == 7:
            return "97"
        if sh == 8:
            return "98"
        return _EXTRA_BUCKET

    if hi == 10:
        if sh <= 5:
            return "T5"
        if sh == 6:
            return "T6"
        if sh == 7:
            return "T7"
        if sh == 8:
            return "T8"
        if sh == 9:
            return "T9"
        return _EXTRA_BUCKET

    if hi == 11:
        return "J8" if sh <= 8 else "J9"

    if hi == 12:
        return "Q"
    if hi == 13:
        return "K"
    if hi == 14:
        return "A"

    return _EXTRA_BUCKET


# bucket logic for the 1-draw abstraction
def bucket_label_1draw(hand: Tuple[int, ...]) -> str:
    cat, _ = classify_27(hand)

    if cat == 4:
        return STRAIGHT_BUCKET

    if cat >= 2:
        return _EXTRA_BUCKET

    if cat == 1:
        pr = next(r for r, cnt in Counter(hand).items() if cnt == 2)
        if 2 <= pr <= 9:
            return f"{pr}{pr}"
        return _EXTRA_BUCKET

    s = sorted(hand)
    hi, sh = s[-1], s[-2]
    return _highcard_bucket_label(hi, sh)


# Chance model
@dataclass(frozen=True)
# helper for this part of the script
class SeedItem:
    seed: Tuple[int, ...]
    draws: int
    weight: float


@dataclass(frozen=True)
# helper for this part of the script
class Matchup:
    btn_bucket: str
    bb_bucket: str
    prob: float
    win_p: float
    tie_p: float
    lose_p: float


_DRAW_CACHE: Dict[Tuple[Tuple[int, ...], int], List[Tuple[Tuple[int, ...], float, Tuple[int, ...]]]] = {}


# lists all possible draws from the remaining rank deck with their weights
def _draw_outcomes(counts: Tuple[int, ...], k: int) -> List[Tuple[Tuple[int, ...], float, Tuple[int, ...]]]:
    key = (counts, k)
    cached = _DRAW_CACHE.get(key)
    if cached is not None:
        return cached

    n = sum(counts)
    if k < 0 or k > n:
        _DRAW_CACHE[key] = []
        return _DRAW_CACHE[key]

    if k == 0:
        out = [((), 1.0, counts)]
        _DRAW_CACHE[key] = out
        return out

    if k == 1:
        denom = float(n)
        out = []
        for i, c in enumerate(counts):
            if c <= 0:
                continue
            r = _IDX_TO_RANK[i]
            p = float(c) / denom
            nxt = list(counts)
            nxt[i] -= 1
            out.append(((r,), p, tuple(nxt)))
        _DRAW_CACHE[key] = out
        return out

    if k == 2:
        denom = comb(n, 2)
        if denom <= 0:
            _DRAW_CACHE[key] = []
            return _DRAW_CACHE[key]
        denom = float(denom)
        out = []

        for i, c in enumerate(counts):
            if c >= 2:
                ways = comb(c, 2)
                p = float(ways) / denom
                r = _IDX_TO_RANK[i]
                nxt = list(counts)
                nxt[i] -= 2
                out.append(((r, r), p, tuple(nxt)))

        for i, ci in enumerate(counts):
            if ci <= 0:
                continue
            for j in range(i + 1, len(counts)):
                cj = counts[j]
                if cj <= 0:
                    continue
                ways = ci * cj
                p = float(ways) / denom
                r1 = _IDX_TO_RANK[i]
                r2 = _IDX_TO_RANK[j]
                nxt = list(counts)
                nxt[i] -= 1
                nxt[j] -= 1
                out.append(((r1, r2), p, tuple(nxt)))

        _DRAW_CACHE[key] = out
        return out

    raise ValueError(f"Unsupported k={k} (only 0/1/2 supported).")


# SIM root bucket frequencies
def _sample_index_weighted(w: np.ndarray) -> int:
    w = np.asarray(w, dtype=float)
    s = float(w.sum())
    if s <= 0:
        return 0
    r = random.random() * s
    acc = 0.0
    for i, wi in enumerate(w):
        acc += float(wi)
        if r <= acc:
            return i
    return len(w) - 1


# samples one rank from the remaining deck counts
def _sample_one_from_counts(counts: List[int]) -> int:
    n = sum(counts)
    if n <= 0:
        raise RuntimeError("empty deck")
    r = random.randrange(n)
    acc = 0
    for i, c in enumerate(counts):
        acc += c
        if r < acc:
            counts[i] -= 1
            return i
    i = len(counts) - 1
    counts[i] -= 1
    return i


# monte carlo estimate of the starting bucket frequencies
def estimate_bucket_freqs_sim_seedmodel(
    n: int,
    btn_items: List[SeedItem],
    bb_items: List[SeedItem],
    buckets: List[str],
    bucket_label_fn: Callable[[Tuple[int, ...]], str],
) -> Dict[str, Dict[str, float]]:
    btn_ct = Counter()
    bb_ct = Counter()

    btn_w = np.array([float(it.weight) for it in btn_items], dtype=float)
    bb_w = np.array([float(it.weight) for it in bb_items], dtype=float)

    buckets = [b for b in buckets if b != _EXTRA_BUCKET]

    for _ in range(int(n)):
        bi = btn_items[_sample_index_weighted(btn_w)]
        oi = bb_items[_sample_index_weighted(bb_w)]

        counts0 = _counts_after_seeds(bi.seed, oi.seed)
        if counts0 is None:
            continue

        counts = list(counts0)

        btn_draw: List[int] = []
        ok = True
        for _k in range(int(bi.draws)):
            try:
                idx = _sample_one_from_counts(counts)
            except RuntimeError:
                ok = False
                break
            btn_draw.append(_IDX_TO_RANK[idx])
        if not ok:
            continue

        bb_draw: List[int] = []
        for _k in range(int(oi.draws)):
            try:
                idx = _sample_one_from_counts(counts)
            except RuntimeError:
                ok = False
                break
            bb_draw.append(_IDX_TO_RANK[idx])
        if not ok:
            continue

        btn_hand = tuple(sorted(bi.seed + tuple(btn_draw)))
        bb_hand = tuple(sorted(oi.seed + tuple(bb_draw)))

        b0 = bucket_label_fn(btn_hand)
        b1 = bucket_label_fn(bb_hand)

        if b0 != _EXTRA_BUCKET:
            btn_ct[b0] += 1
        if b1 != _EXTRA_BUCKET:
            bb_ct[b1] += 1

    def norm(ct: Counter) -> Dict[str, float]:
        s = float(sum(ct.values()))
        out = {b: 0.0 for b in buckets}
        if s <= 0:
            return out
        for b in buckets:
            out[b] = float(ct.get(b, 0)) / s
        return out

    return {"BTN": norm(btn_ct), "BB": norm(bb_ct)}


# builds the btn vs bb bucket matchup table used by training
def build_bucket_pair_matchups_seedmodel(
    btn_items: List[SeedItem],
    bb_items: List[SeedItem],
    buckets: List[str],
    bucket_label_fn: Callable[[Tuple[int, ...]], str],
    sim_init_freq_n: int,
) -> Tuple[List[Matchup], Dict[str, Dict[str, float]]]:
    b2i = {b: i for i, b in enumerate(buckets)}
    nB = len(buckets)

    joint = np.zeros((nB, nB), dtype=float)
    win = np.zeros((nB, nB), dtype=float)
    tie = np.zeros((nB, nB), dtype=float)
    lose = np.zeros((nB, nB), dtype=float)

    s_btn = float(sum(it.weight for it in btn_items))
    s_bb = float(sum(it.weight for it in bb_items))
    if s_btn <= 0 or s_bb <= 0:
        raise RuntimeError("seed weights sum to zero")

    total_mass = 0.0

    _cls_cache: Dict[Tuple[int, ...], Tuple[int, Tuple[int, ...]]] = {}

    def _classify_cached(h: Tuple[int, ...]) -> Tuple[int, Tuple[int, ...]]:
        v = _cls_cache.get(h)
        if v is not None:
            return v
        v = classify_27(h)
        _cls_cache[h] = v
        return v

    def _compare_cached(h1: Tuple[int, ...], h2: Tuple[int, ...]) -> int:
        c1, v1 = _classify_cached(h1)
        c2, v2 = _classify_cached(h2)
        if c1 < c2:
            return +1
        if c1 > c2:
            return -1
        if v1 < v2:
            return +1
        if v1 > v2:
            return -1
        return 0

    for bi in btn_items:
        p_btn_seed = float(bi.weight) / s_btn
        btn_seed = bi.seed
        for oi in bb_items:
            p_bb_seed = float(oi.weight) / s_bb
            bb_seed = oi.seed

            p_seed = p_btn_seed * p_bb_seed
            if p_seed <= 0:
                continue

            counts0 = _counts_after_seeds(btn_seed, bb_seed)
            if counts0 is None:
                continue

            btn_out = _draw_outcomes(counts0, bi.draws)
            if not btn_out:
                continue

            for d_btn, p_btn, counts1 in btn_out:
                if p_btn <= 0:
                    continue

                btn_hand = tuple(sorted(btn_seed + d_btn))
                b0 = bucket_label_fn(btn_hand)
                if b0 == _EXTRA_BUCKET:
                    continue
                i0 = b2i.get(b0)
                if i0 is None:
                    continue

                bb_out = _draw_outcomes(counts1, oi.draws)
                if not bb_out:
                    continue

                for d_bb, p_bb, _counts2 in bb_out:
                    if p_bb <= 0:
                        continue

                    bb_hand = tuple(sorted(bb_seed + d_bb))
                    b1 = bucket_label_fn(bb_hand)
                    if b1 == _EXTRA_BUCKET:
                        continue
                    i1 = b2i.get(b1)
                    if i1 is None:
                        continue

                    mass = p_seed * p_btn * p_bb
                    total_mass += mass
                    joint[i0, i1] += mass

                    cmpv = _compare_cached(btn_hand, bb_hand)
                    if cmpv > 0:
                        win[i0, i1] += mass
                    elif cmpv < 0:
                        lose[i0, i1] += mass
                    else:
                        tie[i0, i1] += mass

    if total_mass <= 0:
        raise RuntimeError("bucket-pair mass is zero")

    joint /= total_mass
    win /= total_mass
    tie /= total_mass
    lose /= total_mass

    matchups: List[Matchup] = []
    for i0, b0 in enumerate(buckets):
        for i1, b1 in enumerate(buckets):
            p = float(joint[i0, i1])
            if p <= 0:
                continue
            wp = float(win[i0, i1] / p)
            tp = float(tie[i0, i1] / p)
            lp = float(lose[i0, i1] / p)
            matchups.append(Matchup(b0, b1, p, wp, tp, lp))

    bucket_freq_by_player = estimate_bucket_freqs_sim_seedmodel(
        int(sim_init_freq_n),
        btn_items=btn_items,
        bb_items=bb_items,
        buckets=buckets,
        bucket_label_fn=bucket_label_fn,
    )

    return matchups, bucket_freq_by_player


# CFR+ primitives
A_CHECK, A_BET, A_CALL, A_FOLD, A_RAISE = "k", "b", "c", "f", "r"

ACTION_LABELS = {
    A_CHECK: "check",
    A_BET: "bet",
    A_CALL: "call",
    A_FOLD: "fold",
    A_RAISE: "raise",
}


# helper for this part of the script
class InfoSetPlus:
    __slots__ = ("acts", "regrets", "strat_sum", "strat_w")

    def __init__(self, acts: List[str]):
        self.acts = acts[:]
        self.regrets = np.zeros(len(acts), dtype=float)
        self.strat_sum = np.zeros(len(acts), dtype=float)
        self.strat_w = 0.0

    def rm(self) -> np.ndarray:
        pos = np.maximum(self.regrets, 0.0)
        s = float(pos.sum())
        if s <= 0:
            return np.ones(len(pos), dtype=float) / len(pos)
        return pos / s

    def record_avg(self, probs: np.ndarray, weight: float) -> None:
        if weight <= 0:
            return
        self.strat_sum += weight * probs
        self.strat_w += weight

    def avg_strategy(self) -> np.ndarray:
        if self.strat_w <= 0:
            return np.ones(len(self.acts), dtype=float) / len(self.acts)
        out = self.strat_sum / self.strat_w
        s = float(out.sum())
        return out / s if s > 0 else (np.ones(len(out), dtype=float) / len(out))


# helper for this part of the script
def payoff_chip_btn_terminal_bucketgame(
    s,
    m: Matchup,
    root_pot: float,
    root_inv_btn: float,
    root_inv_bb: float,
) -> float:
    dI0 = float(s.invested_btn) - float(root_inv_btn)
    dI1 = float(s.invested_bb) - float(root_inv_bb)

    if getattr(s, "showdown", False):
        win = float(root_pot) + dI1
        lose = -dI0
        tie = 0.5 * float(root_pot) + 0.5 * (dI1 - dI0)
        return float(m.win_p) * win + float(m.lose_p) * lose + float(m.tie_p) * tie

    return (float(root_pot) + dI1) if getattr(s, "winner", None) == 0 else (-dI0)


# helper for this part of the script
def payoff_zero_sum_btn_terminal(
    s,
    m: Matchup,
    root_pot: float,
    root_inv_btn: float,
    root_inv_bb: float,
) -> float:
    return payoff_chip_btn_terminal_bucketgame(s, m, root_pot, root_inv_btn, root_inv_bb) - (float(root_pot) / 2.0)


# helper for this part of the script
def terminal_utility(
    s,
    m: Matchup,
    perspective: int,
    root_pot: float,
    root_inv_btn: float,
    root_inv_bb: float,
) -> float:
    u0 = payoff_zero_sum_btn_terminal(s, m, root_pot, root_inv_btn, root_inv_bb)
    return u0 if perspective == 0 else -u0


infosets: Dict[tuple, InfoSetPlus] = {}


# clears the infosets so a fresh run starts cleanly
def reset_infosets() -> None:
    global infosets
    infosets = {}


# helper for this part of the script
def cfrplus_traverse(
    m: Matchup,
    s,
    r0: float,
    r1: float,
    perspective: int,
    iteration: int,
    root_pot: float,
    root_inv_btn: float,
    root_inv_bb: float,
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
) -> float:
    if getattr(s, "terminal", False):
        return terminal_utility(s, m, perspective, root_pot, root_inv_btn, root_inv_bb)

    p = int(s.to_act)
    acts = legal_actions_fn(s)
    if not acts:
        return 0.0

    b = m.btn_bucket if p == 0 else m.bb_bucket
    k = ikey_fn(p, b, s)

    node = infosets.get(k)
    if node is None:
        node = InfoSetPlus(acts)
        infosets[k] = node

    sigma = node.rm()

    reach_p = r0 if p == 0 else r1
    node.record_avg(sigma, float(iteration) * float(reach_p))

    utils = np.zeros(len(acts), dtype=float)
    node_util = 0.0

    for i, a in enumerate(acts):
        ns = step_fn(s, a)
        if p == 0:
            utils[i] = cfrplus_traverse(
                m, ns, r0 * float(sigma[i]), r1, perspective, iteration,
                root_pot, root_inv_btn, root_inv_bb,
                legal_actions_fn, step_fn, ikey_fn
            )
        else:
            utils[i] = cfrplus_traverse(
                m, ns, r0, r1 * float(sigma[i]), perspective, iteration,
                root_pot, root_inv_btn, root_inv_bb,
                legal_actions_fn, step_fn, ikey_fn
            )
        node_util += float(sigma[i]) * utils[i]

    if p == perspective:
        opp_reach = r1 if p == 0 else r0
        node.regrets += float(opp_reach) * (utils - node_util)
        node.regrets = np.maximum(node.regrets, 0.0)

    return float(node_util)


# small helper used by the main parts of the script
def _avg_probs_for_state(
    p: int,
    b: str,
    s,
    legal_actions_fn: Callable[[object], List[str]],
    ikey_fn: Callable[[int, str, object], tuple],
) -> Tuple[List[str], np.ndarray]:
    acts = legal_actions_fn(s)
    node = infosets.get(ikey_fn(p, b, s))
    if node is None:
        return acts, np.ones(len(acts), dtype=float) / max(1, len(acts))

    avg = node.avg_strategy()
    if len(avg) != len(acts):
        return acts, np.ones(len(acts), dtype=float) / max(1, len(acts))

    return acts, avg


# evaluates the state value from the button perspective
def eval_chip_btn(
    m: Matchup,
    s,
    root_pot: float,
    root_inv_btn: float,
    root_inv_bb: float,
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
) -> float:
    if getattr(s, "terminal", False):
        return payoff_chip_btn_terminal_bucketgame(s, m, root_pot, root_inv_btn, root_inv_bb)

    p = int(s.to_act)
    b = m.btn_bucket if p == 0 else m.bb_bucket
    acts, probs = _avg_probs_for_state(p, b, s, legal_actions_fn, ikey_fn)
    if not acts:
        return 0.0

    total = 0.0
    for i, a in enumerate(acts):
        total += float(probs[i]) * eval_chip_btn(
            m, step_fn(s, a),
            root_pot, root_inv_btn, root_inv_bb,
            legal_actions_fn, step_fn, ikey_fn
        )
    return float(total)


# helper for this part of the script
def eval_ev_btn(
    root_state,
    matchups: List[Matchup],
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
) -> float:
    root_pot = float(root_state.pot)
    root_inv_btn = float(root_state.invested_btn)
    root_inv_bb = float(root_state.invested_bb)
    return sum(
        float(m.prob) * eval_chip_btn(
            m, root_state, root_pot, root_inv_btn, root_inv_bb,
            legal_actions_fn, step_fn, ikey_fn
        )
        for m in matchups
    )


# Node-locked BR / exploitability
def concrete_key(base_key: str, bucket: str) -> str:
    return base_key + "_" + bucket


# builds one part of the data or output pipeline
def build_best_response_policy(action_sums: Dict[str, List[float]]) -> Dict[str, int]:
    pol: Dict[str, int] = {}
    for k, sums in action_sums.items():
        best_i = 0
        best_v = -1e300
        for i, v in enumerate(sums):
            if v > best_v:
                best_v = v
                best_i = i
        pol[k] = best_i
    return pol


# helper for this part of the script
def same_policy(a: Dict[str, int], b: Dict[str, int]) -> bool:
    return len(a) == len(b) and all(b.get(k) == v for k, v in a.items())


# helper for this part of the script
def terminal_chip_target(
    s,
    m: Matchup,
    target: int,
    root_pot: float,
    root_inv_btn: float,
    root_inv_bb: float,
) -> float:
    btn_chip = payoff_chip_btn_terminal_bucketgame(s, m, root_pot, root_inv_btn, root_inv_bb)
    return btn_chip if target == 0 else (float(root_pot) - btn_chip)


# helper for this part of the script
def accumulate_policy_improvement(
    s,
    m: Matchup,
    target: int,
    policy: Dict[str, int],
    action_sums: Dict[str, List[float]],
    reach_weight: float,
    root_pot: float,
    root_inv_btn: float,
    root_inv_bb: float,
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    info_key_str_fn: Callable[[object], str],
) -> float:
    if getattr(s, "terminal", False):
        return terminal_chip_target(s, m, target, root_pot, root_inv_btn, root_inv_bb)

    p = int(s.to_act)
    acts = legal_actions_fn(s)
    if not acts:
        return 0.0

    if p == target:
        my_bucket = m.btn_bucket if target == 0 else m.bb_bucket
        key = concrete_key(info_key_str_fn(s), my_bucket)

        sums = action_sums.get(key)
        if sums is None or len(sums) != len(acts):
            sums = [0.0] * len(acts)

        child_vals = [0.0] * len(acts)
        for i, a in enumerate(acts):
            v = accumulate_policy_improvement(
                step_fn(s, a), m, target, policy, action_sums,
                reach_weight, root_pot, root_inv_btn, root_inv_bb,
                legal_actions_fn, step_fn, ikey_fn, info_key_str_fn
            )
            child_vals[i] = v
            sums[i] += float(reach_weight) * float(v)

        action_sums[key] = sums

        idx = int(policy.get(key, 0))
        if idx < 0 or idx >= len(child_vals):
            idx = 0
        return float(child_vals[idx])

    opp_bucket = m.btn_bucket if p == 0 else m.bb_bucket
    _, probs = _avg_probs_for_state(p, opp_bucket, s, legal_actions_fn, ikey_fn)

    total = 0.0
    for i, a in enumerate(acts):
        pr = float(probs[i])
        total += pr * accumulate_policy_improvement(
            step_fn(s, a), m, target, policy, action_sums,
            float(reach_weight) * pr, root_pot, root_inv_btn, root_inv_bb,
            legal_actions_fn, step_fn, ikey_fn, info_key_str_fn
        )
    return float(total)


# helper for this part of the script
def evaluate_best_response_with_policy(
    s,
    m: Matchup,
    target: int,
    policy: Dict[str, int],
    root_pot: float,
    root_inv_btn: float,
    root_inv_bb: float,
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    info_key_str_fn: Callable[[object], str],
) -> float:
    if getattr(s, "terminal", False):
        return terminal_chip_target(s, m, target, root_pot, root_inv_btn, root_inv_bb)

    p = int(s.to_act)
    acts = legal_actions_fn(s)
    if not acts:
        return 0.0

    if p == target:
        my_bucket = m.btn_bucket if target == 0 else m.bb_bucket
        key = concrete_key(info_key_str_fn(s), my_bucket)
        idx = int(policy.get(key, 0))
        if idx < 0 or idx >= len(acts):
            idx = 0
        return evaluate_best_response_with_policy(
            step_fn(s, acts[idx]), m, target, policy,
            root_pot, root_inv_btn, root_inv_bb,
            legal_actions_fn, step_fn, ikey_fn, info_key_str_fn
        )

    opp_bucket = m.btn_bucket if p == 0 else m.bb_bucket
    _, probs = _avg_probs_for_state(p, opp_bucket, s, legal_actions_fn, ikey_fn)

    total = 0.0
    for i, a in enumerate(acts):
        total += float(probs[i]) * evaluate_best_response_with_policy(
            step_fn(s, a), m, target, policy,
            root_pot, root_inv_btn, root_inv_bb,
            legal_actions_fn, step_fn, ikey_fn, info_key_str_fn
        )
    return float(total)


# helper for this part of the script
def compute_best_response_ev(
    root_state,
    matchups: List[Matchup],
    target: int,
    policy_init: Optional[Dict[str, int]],
    policy_iters: int,
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    info_key_str_fn: Callable[[object], str],
) -> Tuple[float, float, Dict[str, int]]:
    root_pot = float(root_state.pot)
    root_inv_btn = float(root_state.invested_btn)
    root_inv_bb = float(root_state.invested_bb)

    policy: Dict[str, int] = dict(policy_init) if isinstance(policy_init, dict) else {}

    for _ in range(int(policy_iters)):
        action_sums: Dict[str, List[float]] = {}
        for m in matchups:
            accumulate_policy_improvement(
                root_state, m, target, policy, action_sums, float(m.prob),
                root_pot, root_inv_btn, root_inv_bb,
                legal_actions_fn, step_fn, ikey_fn, info_key_str_fn
            )
        nxt = build_best_response_policy(action_sums)
        if same_policy(policy, nxt):
            policy = nxt
            break
        policy = nxt

    total_target = 0.0
    for m in matchups:
        total_target += float(m.prob) * evaluate_best_response_with_policy(
            root_state, m, target, policy,
            root_pot, root_inv_btn, root_inv_bb,
            legal_actions_fn, step_fn, ikey_fn, info_key_str_fn
        )

    if target == 0:
        ev_btn, ev_bb = float(total_target), float(root_pot) - float(total_target)
    else:
        ev_bb, ev_btn = float(total_target), float(root_pot) - float(total_target)

    return ev_btn, ev_bb, policy


# Training (CFR+)
def should_recalc_exploit_default(it: int, total_iters: int) -> bool:
    if it <= 500:
        return True
    if it <= 10_000:
        return (it % 10 == 0)
    return (it % 100 == 0) or (it == total_iters)


# helper for this part of the script
def should_track_regret_default(it: int, total_iters: int) -> bool:
    return (it % 10 == 0) or (it == total_iters)


# main training loop for the full chance sweep cfr+ run
def train_cfrplus_sweep(
    *,
    matchups: List[Matchup],
    buckets: List[str],
    report_seqs: List[str],
    iterations: int,
    pot: float,
    initial_state_fn: Callable[[float], object],
    state_from_history_fn: Callable[[str, float], object],
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    info_key_str_fn: Callable[[object], str],
    br_policy_iters: int,
    big_blind: float,
    exp_print_step: int,
    stop_expl_mbb: Optional[float] = None,
    dense_expl_below_mbb: Optional[float] = None,
) -> Tuple[
    List[int], dict, dict, dict,
    List[int], List[float], List[float], List[float]
]:
    global infosets

    regret_data: Dict[tuple, Dict[str, List[float]]] = {}
    evo_data: Dict[tuple, Dict[str, List[float]]] = {}
    node_actions: Dict[tuple, List[str]] = {}

    for seq in report_seqs:
        s0 = state_from_history_fn(seq, float(pot))
        if getattr(s0, "terminal", False):
            continue
        p = int(s0.to_act)
        acts = legal_actions_fn(s0)
        for bucket in buckets:
            key = ikey_fn(p, bucket, s0)
            regret_data[key] = {a: [] for a in acts}
            evo_data[key] = {a: [] for a in acts}
            node_actions[key] = acts[:]

    tracked_iters: List[int] = []

    exploit_iters: List[int] = []
    expl_mbb: List[float] = []
    expl_bb: List[float] = []
    expl_chip: List[float] = []

    root = initial_state_fn(float(pot))
    root_pot = float(root.pot)
    root_inv_btn = float(root.invested_btn)
    root_inv_bb = float(root.invested_bb)

    br_pol_btn: Dict[str, int] = {}
    br_pol_bb: Dict[str, int] = {}

    last_chip = 0.0
    last_bb = 0.0
    last_mbb = 0.0

    BB_UNIT = float(big_blind)
    MBB_UNIT = float(big_blind) / 1000.0
    print_step = max(int(iterations) // 5, 1)

    stop_target = None if stop_expl_mbb is None else float(stop_expl_mbb)
    dense_target = None if dense_expl_below_mbb is None else float(dense_expl_below_mbb)

    if stop_target is not None and dense_target is None:
        dense_target = stop_target * 1.25 if stop_target > 0 else 0.0

    if stop_target is not None and dense_target is not None and dense_target < stop_target:
        dense_target = stop_target

    dense_expl_mode = False
    stop_hit = False

    t_start = time.perf_counter()
    t_sweep = 0.0
    t_track = 0.0
    t_expl = 0.0

    for it in range(1, int(iterations) + 1):
        t0 = time.perf_counter()
        for m in matchups:
            w = float(m.prob)
            if w <= 0:
                continue
            cfrplus_traverse(
                m, root, w, w, 0, it,
                root_pot, root_inv_btn, root_inv_bb,
                legal_actions_fn, step_fn, ikey_fn
            )
            cfrplus_traverse(
                m, root, w, w, 1, it,
                root_pot, root_inv_btn, root_inv_bb,
                legal_actions_fn, step_fn, ikey_fn
            )
        t_sweep += (time.perf_counter() - t0)

        if should_track_regret_default(it, int(iterations)):
            t1 = time.perf_counter()
            tracked_iters.append(it)
            for key, acts in node_actions.items():
                node = infosets.get(key)
                if node is None:
                    uni = 100.0 / max(1, len(acts))
                    for a in acts:
                        regret_data[key][a].append(0.0)
                        evo_data[key][a].append(round(uni, 4))
                    continue

                rd = {a: float(node.regrets[j]) for j, a in enumerate(node.acts)}
                base = node.rm()
                aidx = {a: j for j, a in enumerate(node.acts)}
                for a in acts:
                    regret_data[key][a].append(round(rd.get(a, 0.0), 6))
                    evo_data[key][a].append(round(100.0 * float(base[aidx[a]]), 4))
            t_track += (time.perf_counter() - t1)

        recalc_expl = dense_expl_mode or should_recalc_exploit_default(it, int(iterations))

        if recalc_expl:
            t2 = time.perf_counter()

            ev_btn = eval_ev_btn(root, matchups, legal_actions_fn, step_fn, ikey_fn)
            ev_bb = float(root_pot) - float(ev_btn)

            br_btn_ev_btn, _, br_pol_btn = compute_best_response_ev(
                root, matchups, 0, br_pol_btn, int(br_policy_iters),
                legal_actions_fn, step_fn, ikey_fn, info_key_str_fn
            )
            _, br_bb_ev_bb, br_pol_bb = compute_best_response_ev(
                root, matchups, 1, br_pol_bb, int(br_policy_iters),
                legal_actions_fn, step_fn, ikey_fn, info_key_str_fn
            )

            delta_btn = max(0.0, float(br_btn_ev_btn) - float(ev_btn))
            delta_bb = max(0.0, float(br_bb_ev_bb) - float(ev_bb))
            total = 0.5 * (delta_btn + delta_bb)

            last_chip = float(total)
            last_bb = float(total / BB_UNIT) if BB_UNIT > 0 else 0.0
            last_mbb = float(total / MBB_UNIT) if MBB_UNIT > 0 else 0.0

            t_expl += (time.perf_counter() - t2)

            if (not dense_expl_mode) and dense_target is not None and last_mbb <= dense_target:
                dense_expl_mode = True
                print(
                    f"[INFO] dense exploit checks enabled @ iter {it}: "
                    f"{last_mbb:.6f} mbb/g <= {dense_target:.6f} mbb/g"
                )

            if stop_target is not None and last_mbb <= stop_target:
                stop_hit = True

        exploit_iters.append(it)
        expl_chip.append(abs(last_chip))
        expl_bb.append(abs(last_bb))
        expl_mbb.append(abs(last_mbb))

        if it % print_step == 0 or it == int(iterations):
            elapsed = time.perf_counter() - t_start
            ips = (float(it) / elapsed) if elapsed > 0 else 0.0
            total_t = max(elapsed, 1e-12)
            ps = 100.0 * t_sweep / total_t
            pt = 100.0 * t_track / total_t
            pe = 100.0 * t_expl / total_t
            print(
                f"iter {it:7d}/{int(iterations):,} nodes={len(infosets):6d} "
                f"expl={last_mbb:9.3f} mbb/g | "
                f"{elapsed:8.1f}s ({ips:6.1f} it/s) | "
                f"sweep {t_sweep:7.1f}s ({ps:5.1f}%) "
                f"track {t_track:7.1f}s ({pt:5.1f}%) "
                f"expl {t_expl:7.1f}s ({pe:5.1f}%)"
            )

        if exp_print_step > 0 and (it % int(exp_print_step) == 0 or it == int(iterations)):
            print(f"  exploit snapshot @ {it}: {last_chip:.6f} chip | {last_bb:.6f} bb/g | {last_mbb:.3f} mbb/g")

        if stop_hit:
            print(
                f"[STOP] exploitability target reached @ iter {it}: "
                f"{last_mbb:.6f} mbb/g <= {stop_target:.6f} mbb/g"
            )
            break

    elapsed = time.perf_counter() - t_start
    print(f"[TIME] total={elapsed:.2f}s sweep={t_sweep:.2f}s track={t_track:.2f}s expl={t_expl:.2f}s")

    return tracked_iters, regret_data, evo_data, node_actions, exploit_iters, expl_mbb, expl_bb, expl_chip


# EV tables
class _PrefixSampler:
    __slots__ = ("items", "prefix", "total")

    def __init__(self, items: List[Matchup], weights: List[float]):
        self.items = items
        prefix: List[float] = []
        s = 0.0
        for w in weights:
            s += float(w)
            prefix.append(s)
        self.prefix = prefix
        self.total = s

    def sample(self) -> Matchup:
        if not self.items or self.total <= 0:
            raise RuntimeError("PrefixSampler has no mass/items")
        x = random.random() * self.total
        idx = bisect.bisect_left(self.prefix, x)
        if idx < 0:
            idx = 0
        if idx >= len(self.items):
            idx = len(self.items) - 1
        return self.items[idx]


# builds samplers for each bucket so later ev estimates are faster
def _build_conditional_matchup_samplers(
    matchups: List[Matchup],
    buckets: List[str],
    matchup_weights: Optional[List[float]] = None,
) -> Dict[Tuple[int, str], Optional[_PrefixSampler]]:
    by_key: Dict[Tuple[int, str], List[Matchup]] = {}
    w_key: Dict[Tuple[int, str], List[float]] = {}

    for b in buckets:
        by_key[(0, b)] = []
        w_key[(0, b)] = []
        by_key[(1, b)] = []
        w_key[(1, b)] = []

    if matchup_weights is not None and len(matchup_weights) != len(matchups):
        raise ValueError("matchup_weights must be None or the same length as matchups")

    for i, m in enumerate(matchups):
        w = float(matchup_weights[i]) if matchup_weights is not None else float(m.prob)
        if w <= 0.0:
            continue
        by_key[(0, m.btn_bucket)].append(m)
        w_key[(0, m.btn_bucket)].append(w)
        by_key[(1, m.bb_bucket)].append(m)
        w_key[(1, m.bb_bucket)].append(w)

    out: Dict[Tuple[int, str], Optional[_PrefixSampler]] = {}
    for k in by_key.keys():
        items = by_key[k]
        ws = w_key[k]
        if not items or sum(ws) <= 0.0:
            out[k] = None
            continue
        out[k] = _PrefixSampler(items, ws)
    return out


# helper for this part of the script
def compute_ev_rows_for_state(
    s0,
    *,
    pot: float,
    buckets: List[str],
    bucket_freq: Dict[str, float],
    matchups: List[Matchup],
    samples_per_bucket: int,
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    matchup_weights: Optional[List[float]] = None,
) -> dict:
    root_pot = float(s0.pot)
    root_inv_btn = float(s0.invested_btn)
    root_inv_bb = float(s0.invested_bb)

    actor = int(s0.to_act)
    actor_name = "BTN" if actor == 0 else "BB"

    samplers = _build_conditional_matchup_samplers(matchups, buckets, matchup_weights=matchup_weights)

    rows = []
    for b in buckets:
        samp = samplers.get((actor, b))
        rate = round(100.0 * float(bucket_freq.get(b, 0.0)), 6)
        if samp is None:
            rows.append({"bucket": b, "rate": rate, "btn_ev": None, "bb_ev": None, "n": 0})
            continue

        total = 0.0
        for _ in range(int(samples_per_bucket)):
            m = samp.sample()
            total += eval_chip_btn(m, s0, root_pot, root_inv_btn, root_inv_bb, legal_actions_fn, step_fn, ikey_fn)

        btn_chip = total / float(samples_per_bucket)
        bb_chip = float(root_pot) - float(btn_chip)

        rows.append({
            "bucket": b,
            "rate": rate,
            "btn_ev": round(float(btn_chip), 6),
            "bb_ev": round(float(bb_chip), 6),
            "n": int(samples_per_bucket),
        })

    return {
        "history": getattr(s0, "history", ""),
        "actor": actor_name,
        "pot": float(s0.pot),
        "to_call": float(getattr(s0, "to_call", 0.0)),
        "raises_made": int(getattr(s0, "raises_made", 0)),
        "rows": rows,
    }


# keeps excel sheet names valid and short enough
def _safe_sheet_title(title: str, used: set) -> str:
    bad = set(r'[]:*?/\ ')
    cleaned = "".join("_" if ch in bad else ch for ch in title)
    cleaned = cleaned if cleaned else "root"
    cleaned = cleaned[:31]

    base = cleaned
    n = 1
    while cleaned in used:
        suffix = f"_{n}"
        cleaned = (base[:31 - len(suffix)] + suffix)
        n += 1
    used.add(cleaned)
    return cleaned


# gives the excel header row a cleaner look
def _style_ws_header(ws, row_idx: int, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


# just makes the excel columns fit the content better
def _autosize_ws(ws, min_w: int = 10, max_w: int = 18) -> None:
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, min_w), len(val) + 2)

    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_w, min(max_w, w))


# estimates the ev for each action at one state by sampling matchups
def compute_action_ev_rows_for_state(
    s0,
    *,
    buckets: List[str],
    bucket_freq: Dict[str, float],
    matchups: List[Matchup],
    samples_per_bucket: int,
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    matchup_weights: Optional[List[float]] = None,
) -> dict:
    root_pot = float(s0.pot)
    root_inv_btn = float(s0.invested_btn)
    root_inv_bb = float(s0.invested_bb)

    actor = int(s0.to_act)
    actor_name = "BTN" if actor == 0 else "BB"
    acts = legal_actions_fn(s0)

    samplers = _build_conditional_matchup_samplers(
        matchups,
        buckets,
        matchup_weights=matchup_weights,
    )

    rows = []

    total_rate = 0.0
    overall_node_btn_num = 0.0
    overall_action_btn_num = {a: 0.0 for a in acts}

    for b in buckets:
        rate = float(bucket_freq.get(b, 0.0))
        rate_pct = 100.0 * rate
        samp = samplers.get((actor, b))

        row = {
            "bucket": b,
            "rate": round(rate_pct, 6),
            "btn_ev": None,
            "bb_ev": None,
        }

        for a in acts:
            row[f"{a}_btn_ev"] = None
            row[f"{a}_bb_ev"] = None

        if samp is None or rate <= 0.0:
            rows.append(row)
            continue

        total_node_btn = 0.0
        action_btn_totals = {a: 0.0 for a in acts}

        for _ in range(int(samples_per_bucket)):
            m = samp.sample()

            total_node_btn += eval_chip_btn(
                m, s0, root_pot, root_inv_btn, root_inv_bb,
                legal_actions_fn, step_fn, ikey_fn
            )

            for a in acts:
                ns = step_fn(s0, a)
                action_btn_totals[a] += eval_chip_btn(
                    m, ns, root_pot, root_inv_btn, root_inv_bb,
                    legal_actions_fn, step_fn, ikey_fn
                )

        btn_ev = total_node_btn / float(samples_per_bucket)
        bb_ev = float(root_pot) - float(btn_ev)

        row["btn_ev"] = round(float(btn_ev), 6)
        row["bb_ev"] = round(float(bb_ev), 6)

        total_rate += rate
        overall_node_btn_num += rate * btn_ev

        for a in acts:
            a_btn = action_btn_totals[a] / float(samples_per_bucket)
            a_bb = float(root_pot) - float(a_btn)

            row[f"{a}_btn_ev"] = round(float(a_btn), 6)
            row[f"{a}_bb_ev"] = round(float(a_bb), 6)

            overall_action_btn_num[a] += rate * a_btn

        rows.append(row)

    overall = {
        "actor": actor_name,
        "history": getattr(s0, "history", ""),
        "pot": float(s0.pot),
        "to_call": float(getattr(s0, "to_call", 0.0)),
        "raises_made": int(getattr(s0, "raises_made", 0)),
        "reach_freq_pct": round(100.0 * total_rate, 6),
        "btn_ev": None,
        "bb_ev": None,
        "actions": acts[:],
    }

    if total_rate > 0.0:
        overall_btn = overall_node_btn_num / total_rate
        overall["btn_ev"] = round(float(overall_btn), 6)
        overall["bb_ev"] = round(float(root_pot - overall_btn), 6)

        for a in acts:
            a_btn = overall_action_btn_num[a] / total_rate
            overall[f"{a}_btn_ev"] = round(float(a_btn), 6)
            overall[f"{a}_bb_ev"] = round(float(root_pot - a_btn), 6)
    else:
        for a in acts:
            overall[f"{a}_btn_ev"] = None
            overall[f"{a}_bb_ev"] = None

    return {
        "actor": actor_name,
        "actions": acts[:],
        "rows": rows,
        "overall": overall,
    }


# writes this part of the output to file
def export_move_evs_excel(
    *,
    out_prefix: str,
    pot: float,
    report_seqs: List[str],
    buckets: List[str],
    matchups: List[Matchup],
    seq_info: Dict[str, dict],
    reach_prob: Dict[str, List[float]],
    state_from_history_fn: Callable[[str, float], object],
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    order_actions_fn: Callable[[List[str]], List[str]],
    ev_samples_per_bucket: int,
) -> None:
    out_dir = os.path.join("..", "data", str(out_prefix), "excel")
    os.makedirs(out_dir, exist_ok=True)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    used_titles = {"Summary"}

    summary_rows = []

    all_actions = [A_CHECK, A_BET, A_CALL, A_FOLD, A_RAISE]

    for seq in report_seqs:
        s0 = state_from_history_fn(seq, float(pot))
        seq_label = seq if seq else "root"

        if getattr(s0, "terminal", False):
            summary_rows.append({
                "sequence": seq_label,
                "actor": "terminal",
                "reach_freq_pct": None,
                "btn_ev": None,
                "bb_ev": None,
            })
            continue

        pr_list = reach_prob.get(seq, [0.0] * len(matchups))
        w_hist = [float(m.prob) * float(pr_list[i]) for i, m in enumerate(matchups)]
        freq_map = seq_info.get(seq, {}).get("bucket_freq", {})

        pack = compute_action_ev_rows_for_state(
            s0,
            buckets=buckets,
            bucket_freq=freq_map,
            matchups=matchups,
            samples_per_bucket=int(ev_samples_per_bucket),
            legal_actions_fn=legal_actions_fn,
            step_fn=step_fn,
            ikey_fn=ikey_fn,
            matchup_weights=w_hist,
        )

        ordered_actions = order_actions_fn(pack["actions"])

        sheet_name = _safe_sheet_title(f"seq_{seq_label}", used_titles)
        ws = wb.create_sheet(title=sheet_name)

        ws["A1"] = "Sequence"
        ws["B1"] = seq_label
        ws["A2"] = "Actor"
        ws["B2"] = pack["overall"]["actor"]
        ws["A3"] = "Pot"
        ws["B3"] = pack["overall"]["pot"]
        ws["D1"] = "To Call"
        ws["E1"] = pack["overall"]["to_call"]
        ws["D2"] = "Raises Made"
        ws["E2"] = pack["overall"]["raises_made"]
        ws["D3"] = "Reach Freq %"
        ws["E3"] = pack["overall"]["reach_freq_pct"]

        header = ["Bucket", "Rate %", "BTN EV", "BB EV"]
        for a in ordered_actions:
            label = ACTION_LABELS[a]
            header.extend([f"{label} BTN EV", f"{label} BB EV"])

        start_row = 5
        for j, h in enumerate(header, start=1):
            ws.cell(row=start_row, column=j, value=h)
        _style_ws_header(ws, start_row, len(header))

        r = start_row + 1
        for row in pack["rows"]:
            out_vals = [row["bucket"], row["rate"], row["btn_ev"], row["bb_ev"]]
            for a in ordered_actions:
                out_vals.extend([row[f"{a}_btn_ev"], row[f"{a}_bb_ev"]])

            for c, v in enumerate(out_vals, start=1):
                ws.cell(row=r, column=c, value=v)
            r += 1

        r += 1
        ws.cell(row=r, column=1, value="OVERALL")
        ws.cell(row=r, column=2, value=pack["overall"]["reach_freq_pct"])
        ws.cell(row=r, column=3, value=pack["overall"]["btn_ev"])
        ws.cell(row=r, column=4, value=pack["overall"]["bb_ev"])

        c = 5
        for a in ordered_actions:
            ws.cell(row=r, column=c, value=pack["overall"][f"{a}_btn_ev"])
            ws.cell(row=r, column=c + 1, value=pack["overall"][f"{a}_bb_ev"])
            c += 2

        for c in range(1, len(header) + 1):
            ws.cell(row=r, column=c).font = Font(bold=True)

        _autosize_ws(ws)

        sm = {
            "sequence": seq_label,
            "actor": pack["overall"]["actor"],
            "reach_freq_pct": pack["overall"]["reach_freq_pct"],
            "btn_ev": pack["overall"]["btn_ev"],
            "bb_ev": pack["overall"]["bb_ev"],
        }
        for a in all_actions:
            sm[f"{a}_btn_ev"] = pack["overall"].get(f"{a}_btn_ev")
            sm[f"{a}_bb_ev"] = pack["overall"].get(f"{a}_bb_ev")
        summary_rows.append(sm)

    summary_header = ["Sequence", "Actor", "Reach Freq %", "BTN EV", "BB EV"]
    for a in all_actions:
        label = ACTION_LABELS[a]
        summary_header.extend([f"{label} BTN EV", f"{label} BB EV"])

    for j, h in enumerate(summary_header, start=1):
        ws_summary.cell(row=1, column=j, value=h)
    _style_ws_header(ws_summary, 1, len(summary_header))

    rr = 2
    for row in summary_rows:
        vals = [
            row.get("sequence"),
            row.get("actor"),
            row.get("reach_freq_pct"),
            row.get("btn_ev"),
            row.get("bb_ev"),
        ]
        for a in all_actions:
            vals.extend([row.get(f"{a}_btn_ev"), row.get(f"{a}_bb_ev")])

        for c, v in enumerate(vals, start=1):
            ws_summary.cell(row=rr, column=c, value=v)
        rr += 1

    _autosize_ws(ws_summary, max_w=20)

    out_file = os.path.join(out_dir, f"move_evs_pot{int(pot)}.xlsx")
    wb.save(out_file)
    print(f"excel export -> {out_file}")


# Export helpers
def order_actions_limit(acts: List[str]) -> List[str]:
    if "c" in acts:
        return [a for a in ["c", "f", "r"] if a in acts]
    return [a for a in ["k", "b"] if a in acts]


# helper for this part of the script
def compute_bucket_freq_by_sequence(
    *,
    matchups: List[Matchup],
    buckets: List[str],
    report_seqs: List[str],
    pot: float,
    bucket_freq_by_player: Dict[str, Dict[str, float]],
    initial_state_fn: Callable[[float], object],
    state_from_history_fn: Callable[[str, float], object],
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
) -> Tuple[Dict[str, dict], Dict[str, List[float]]]:
    report_set = set(report_seqs)

    seq_info: Dict[str, dict] = {}
    reach_prob: Dict[str, List[float]] = {seq: [0.0] * len(matchups) for seq in report_seqs}

    for seq in report_seqs:
        s0 = state_from_history_fn(seq, float(pot))
        if getattr(s0, "terminal", False):
            seq_info[seq] = {"terminal": True}
            continue
        actor = "BTN" if int(s0.to_act) == 0 else "BB"
        seq_info[seq] = {
            "actor": actor,
            "bucket_freq": {b: 0.0 for b in buckets},
            "bucket_reach": {b: 0.0 for b in buckets},
            "total_freq": 0.0,
        }

    for i, m in enumerate(matchups):
        root = initial_state_fn(float(pot))

        def walk(s, pr: float) -> None:
            if getattr(s, "terminal", False):
                return

            h = getattr(s, "history", "")
            if h in report_set:
                info = seq_info.get(h)
                if info is not None and not info.get("terminal", False):
                    reach_prob[h][i] += float(pr)
                    actor = info["actor"]
                    b = m.btn_bucket if actor == "BTN" else m.bb_bucket
                    info["bucket_freq"][b] = float(info["bucket_freq"].get(b, 0.0)) + float(m.prob) * float(pr)
                    info["total_freq"] = float(info["total_freq"]) + float(m.prob) * float(pr)

            p = int(s.to_act)
            bucket = m.btn_bucket if p == 0 else m.bb_bucket
            acts = legal_actions_fn(s)
            node = infosets.get(ikey_fn(p, bucket, s))
            if node is None:
                probs = [1.0 / max(1, len(acts))] * len(acts)
            else:
                avg = node.avg_strategy()
                probs = list(avg) if len(avg) == len(acts) else [1.0 / max(1, len(acts))] * len(acts)

            for a, pa in zip(acts, probs):
                pa = float(pa)
                if pa <= 0.0:
                    continue
                walk(step_fn(s, a), float(pr) * pa)

        walk(root, 1.0)

    for seq, info in seq_info.items():
        if info.get("terminal", False):
            continue
        actor = info["actor"]
        marg = bucket_freq_by_player.get(actor, {})
        for b in buckets:
            denom = float(marg.get(b, 0.0))
            num = float(info["bucket_freq"].get(b, 0.0))
            info["bucket_reach"][b] = (num / denom) if denom > 0 else 0.0

    return seq_info, reach_prob


# exports the json files that the plots and frontend use
def export_variant(
    *,
    out_prefix: str,
    pot: float,
    iterations: int,
    buckets: List[str],
    report_seqs: List[str],
    bucket_freq_by_player: Dict[str, Dict[str, float]],
    tracked_iters: List[int],
    regret_data: dict,
    evo_data: dict,
    node_actions: dict,
    exploit_iters: List[int],
    expl_mbb: List[float],
    expl_bb: List[float],
    expl_chip: List[float],
    initial_state_fn: Callable[[float], object],
    state_from_history_fn: Callable[[str, float], object],
    legal_actions_fn: Callable[[object], List[str]],
    step_fn: Callable[[object, str], object],
    ikey_fn: Callable[[int, str, object], tuple],
    order_actions_fn: Callable[[List[str]], List[str]],
    matchups: List[Matchup],
    ev_samples_per_bucket: int,
    time_export: bool = False,
) -> None:
    t0 = time.perf_counter() if time_export else 0.0

    out = f"../data/{out_prefix}"
    os.makedirs(out, exist_ok=True)

    buckets = [b for b in buckets if b != _EXTRA_BUCKET]

    with open(f"{out}/bucket_freq_by_player.json", "w", encoding="utf-8") as f:
        json.dump(bucket_freq_by_player, f, indent=2)

    seq_info, reach_prob = compute_bucket_freq_by_sequence(
        matchups=matchups,
        buckets=buckets,
        report_seqs=report_seqs,
        pot=float(pot),
        bucket_freq_by_player=bucket_freq_by_player,
        initial_state_fn=initial_state_fn,
        state_from_history_fn=state_from_history_fn,
        legal_actions_fn=legal_actions_fn,
        step_fn=step_fn,
        ikey_fn=ikey_fn,
    )

    freq_out = {
        "pot": float(pot),
        "bucket_freq_by_player": bucket_freq_by_player,
        "sequences": seq_info,
    }
    with open(f"{out}/bucket_freq_by_sequence_pot{int(pot)}.json", "w", encoding="utf-8") as f:
        json.dump(freq_out, f, indent=2)

    strat_out = {
        "pot": float(pot),
        "iterations": int(iterations),
        "bucket_freq_by_player": bucket_freq_by_player,
        "sequences": {},
    }

    for seq in report_seqs:
        s0 = state_from_history_fn(seq, float(pot))
        if getattr(s0, "terminal", False):
            strat_out["sequences"][seq] = {"terminal": True}
            continue

        p = int(s0.to_act)
        actor = "BTN" if p == 0 else "BB"
        ordered = order_actions_fn(legal_actions_fn(s0))
        freq_map = seq_info.get(seq, {}).get("bucket_freq", {})

        rows = []
        for b in buckets:
            node = infosets.get(ikey_fn(p, b, s0))
            rate = round(100.0 * float(freq_map.get(b, 0.0)), 6)
            if node is None:
                uni = {a: int(round(100 / max(1, len(ordered)))) for a in ordered}
                rows.append({"bucket": b, "rate": rate, **uni})
                continue

            avg = node.avg_strategy()
            am = {a: float(avg[i]) for i, a in enumerate(node.acts)}
            row = {"bucket": b, "rate": rate}
            for a in ordered:
                row[a] = int(round(100.0 * am.get(a, 0.0)))
            rows.append(row)

        sum_rate = sum(float(r.get("rate", 0.0)) for r in rows)
        weighted = {a: 0.0 for a in ordered}
        for row in rows:
            r = float(row.get("rate", 0.0))
            for a in ordered:
                weighted[a] += r * (float(row.get(a, 0.0)) / 100.0)

        if sum_rate > 0.0:
            overall = {a: round((weighted[a] / sum_rate) * 100.0, 4) for a in ordered}
        else:
            overall = {a: 0.0 for a in ordered}

        strat_out["sequences"][seq] = {
            "actor": actor,
            "actions": ordered,
            "overall": overall,
            "rows": rows,
        }

    with open(f"{out}/strategies_pot{int(pot)}.json", "w", encoding="utf-8") as f:
        json.dump(strat_out, f, indent=2)

    reg_out = {"pot": float(pot), "tracked_iterations": tracked_iters, "sequences": {}}
    evo_out = {"pot": float(pot), "tracked_iterations": tracked_iters, "sequences": {}}

    for seq in report_seqs:
        s0 = state_from_history_fn(seq, float(pot))
        if getattr(s0, "terminal", False):
            continue
        p = int(s0.to_act)
        actor = "BTN" if p == 0 else "BB"
        acts = legal_actions_fn(s0)

        reg_out["sequences"][seq] = {"actor": actor, "actions": acts, "buckets": {}}
        evo_out["sequences"][seq] = {"actor": actor, "actions": acts, "buckets": {}}

        for b in buckets:
            key = ikey_fn(p, b, s0)
            reg_out["sequences"][seq]["buckets"][b] = {a: regret_data[key][a] for a in acts}
            evo_out["sequences"][seq]["buckets"][b] = {a: evo_data[key][a] for a in acts}

    with open(f"{out}/regrets_pot{int(pot)}.json", "w", encoding="utf-8") as f:
        json.dump(reg_out, f)

    with open(f"{out}/evolution_pot{int(pot)}.json", "w", encoding="utf-8") as f:
        json.dump(evo_out, f, indent=2)

    ex_out = {
        "tracked_iterations": exploit_iters,
        "exploitability": expl_mbb,
        "exploitability_2": expl_bb,
        "exploitability_3": expl_chip,
    }
    with open(f"{out}/exploitability_pot{int(pot)}.json", "w", encoding="utf-8") as f:
        json.dump(ex_out, f, indent=2)

    ev_out = {"pot": float(pot), "samples_per_bucket": int(ev_samples_per_bucket), "sequences": {}}
    for seq in report_seqs:
        s0 = state_from_history_fn(seq, float(pot))
        if getattr(s0, "terminal", False):
            ev_out["sequences"][seq] = {"terminal": True}
            continue

        pr_list = reach_prob.get(seq, [0.0] * len(matchups))
        w_hist = [float(m.prob) * float(pr_list[i]) for i, m in enumerate(matchups)]
        freq_map = seq_info.get(seq, {}).get("bucket_freq", {})

        ev_out["sequences"][seq] = compute_ev_rows_for_state(
            s0,
            pot=float(pot),
            buckets=buckets,
            bucket_freq=freq_map,
            matchups=matchups,
            samples_per_bucket=int(ev_samples_per_bucket),
            legal_actions_fn=legal_actions_fn,
            step_fn=step_fn,
            ikey_fn=ikey_fn,
            matchup_weights=w_hist,
        )

    with open(f"{out}/ev_pot{int(pot)}.json", "w", encoding="utf-8") as f:
        json.dump(ev_out, f, indent=2)

    export_move_evs_excel(
        out_prefix=str(out_prefix),
        pot=float(pot),
        report_seqs=report_seqs,
        buckets=buckets,
        matchups=matchups,
        seq_info=seq_info,
        reach_prob=reach_prob,
        state_from_history_fn=state_from_history_fn,
        legal_actions_fn=legal_actions_fn,
        step_fn=step_fn,
        ikey_fn=ikey_fn,
        order_actions_fn=order_actions_fn,
        ev_samples_per_bucket=int(ev_samples_per_bucket),
    )

    if time_export:
        print(f"exported {out_prefix} pot{int(pot)} -> {out}/ | export_time={time.perf_counter() - t0:.2f}s")
    else:
        print(f"exported {out_prefix} pot{int(pot)} -> {out}/")


# Limit betting game definition
SEED = 42
ITERATIONS = 50_000
POT_SIZES = [5]

STARTING_PLAYER = 1

MAX_RAISES = 3
RAISE_SCHEDULE = [1.0, 2.0, 3.0, 4.0]

REPORT_SEQS = ["", "k", "b", "kb", "kbr", "br", "brr", "brrr", "kbrr", "kbrrr"]

BIG_BLIND = 0.5
BR_POLICY_ITERS = 16
EXP_PRINT_STEP = 100
EV_SAMPLES_PER_BUCKET = 800
SIM_INIT_FREQ_N = 200_000

btn_grid = [
    ["2345", "2346", "2347", "2348", "2349"],
    ["2345", "2356", "2357", "2358", "2359"],
    ["2345", "2456", "2457", "2458", "2459"],
    ["2345", "3456", "3457", "2458", "3459"],
    ["2346", "2356", "2367", "2368", "2369"],
    ["2346", "2456", "2467", "2468", "2469"],
    ["2346", "3456", "3467", "3468", "3469"],
    ["2356", "2456", "2567", "2568", "2569"],
    ["2356", "3456", "3567", "3568", "3569"],
    ["2347", "2357", "2367", "2378", "2379"],
    ["2347", "2457", "2467", "2478", "2479"],
    ["2347", "3457", "3467", "3478", "3479"],
    ["2357", "2457", "2567", "2578", "2579"],
    ["2357", "3457", "3567", "3578", "3579"],
    ["2457", "3457", "4567", "4578", "4579"],
    ["2367", "2467", "2567", "2678", "2679"],
    ["2367", "3467", "3567", "3678", "3679"],
    ["2348", "2358", "2368", "2378", "2389"],
    ["2348", "2458", "2468", "2478", "2489"],
    ["2348", "3458", "3468", "3478", "3489"],
    ["2358", "2458", "2568", "2578", "2589"],
    ["2358", "3458", "3568", "3578", "3589"],
    ["2458", "3458", "4568", "4578", "4589"],
    ["2368", "2468", "2568", "2678", "2689"],
    ["2368", "3468", "3568", "3678", "3689"],
    ["2468", "3468", "4568", "4678", "4689"],
    ["2568", "3568", "4568", "5678", "5689"],
]

bb_grid = [
    ["2345", "3456", "3457", "2458", "3459"],
    ["2346", "3456", "3467", "3468", "3469"],
    ["2356", "2456", "2567", "2568", "2569"],
    ["2356", "3456", "3567", "3568", "3569"],
    ["2357", "3457", "3567", "3578", "3579"],
    ["2457", "3457", "4567", "4578", "4579"],
    ["2367", "2467", "2567", "2678", "2679"],
    ["2367", "3467", "3567", "3678", "3679"],
    ["2348", "2358", "2368", "2378", "2389"],
    ["2348", "2458", "2468", "2478", "2489"],
    ["2348", "3458", "3468", "3478", "3489"],
    ["2358", "2458", "2568", "2578", "2589"],
    ["2358", "3458", "3568", "3578", "3589"],
    ["2458", "3458", "4568", "4578", "4589"],
    ["2368", "2468", "2568", "2678", "2689"],
    ["2368", "3468", "3568", "3678", "3689"],
    ["2468", "3468", "4568", "4678", "4689"],
    ["2568", "3568", "4568", "5678", "5689"],
]


@dataclass
# helper for this part of the script
class State:
    to_act: int
    history: str
    pot: float
    to_call: float
    raises_made: int
    terminal: bool
    winner: Optional[int]
    showdown: bool
    invested_btn: float
    invested_bb: float


# creates the starting state for the no-limit tree
def initial_state(start_pot: float) -> State:
    return State(
        to_act=STARTING_PLAYER,
        history="",
        pot=float(start_pot),
        to_call=0.0,
        raises_made=0,
        terminal=False,
        winner=None,
        showdown=False,
        invested_btn=0.0,
        invested_bb=0.0,
    )


# returns the legal actions for the current state
def legal_actions(s: State) -> List[str]:
    if s.terminal:
        return []
    if s.to_call == 0:
        return [A_CHECK, A_BET]
    acts = [A_FOLD, A_CALL]
    if s.raises_made < MAX_RAISES:
        acts.append(A_RAISE)
    return acts


# small helper used by the main parts of the script
def _add_invested(st: State, player: int, amt: float) -> None:
    if player == 0:
        st.invested_btn += amt
    else:
        st.invested_bb += amt


# helper for this part of the script
def step(s: State, a: str) -> State:
    t = State(**vars(s))
    t.history += a
    p, opp = s.to_act, 1 - s.to_act

    if s.to_call == 0:
        if a == A_CHECK:
            if t.history.endswith("kk"):
                t.terminal = True
                t.showdown = True
            else:
                t.to_act = opp
        elif a == A_BET:
            amt = float(RAISE_SCHEDULE[0])
            t.pot += amt
            _add_invested(t, p, amt)
            t.to_call = amt
            t.raises_made = 0
            t.to_act = opp
    else:
        if a == A_FOLD:
            t.terminal = True
            t.winner = opp
        elif a == A_CALL:
            t.pot += s.to_call
            _add_invested(t, p, s.to_call)
            t.to_call = 0.0
            t.terminal = True
            t.showdown = True
        elif a == A_RAISE:
            newlvl = float(RAISE_SCHEDULE[s.raises_made + 1])
            curr = t.invested_btn if p == 0 else t.invested_bb
            oppi = t.invested_bb if p == 0 else t.invested_btn
            add = newlvl - curr
            t.pot += add
            _add_invested(t, p, add)
            t.to_call = newlvl - oppi
            t.raises_made = s.raises_made + 1
            t.to_act = opp

    return t


# helper for this part of the script
def state_from_history(h: str, start_pot: float) -> State:
    s = initial_state(start_pot)
    for ch in h:
        s = step(s, ch)
    return s


# helper for this part of the script
def ikey(p: int, bucket: str, s: State) -> tuple:
    return (p, bucket, s.history, s.to_call, s.raises_made)


# helper for this part of the script
def info_key_str(s: State) -> str:
    return f"h:{s.history}|tc:{s.to_call}|rm:{s.raises_made}"


# main
def main():
    global SEED, ITERATIONS, BR_POLICY_ITERS, EXP_PRINT_STEP, EV_SAMPLES_PER_BUCKET, SIM_INIT_FREQ_N

    ap = argparse.ArgumentParser(description="CFR+ 1draw (FULL chance sweep bucket-game)")
    ap.add_argument("--seed", type=int, default=SEED)
    ap.add_argument("--iters", "--iterations", dest="iters", type=int, default=ITERATIONS)
    ap.add_argument("--pots", type=str, default=",".join(str(x) for x in POT_SIZES))
    ap.add_argument("--out-prefix", type=str, default="1draw")
    ap.add_argument("--br-policy-iters", type=int, default=BR_POLICY_ITERS)
    ap.add_argument("--exp-print-step", type=int, default=EXP_PRINT_STEP)
    ap.add_argument("--ev-samples", type=int, default=EV_SAMPLES_PER_BUCKET)
    ap.add_argument("--sim-init-freq-n", type=int, default=SIM_INIT_FREQ_N, help="MC sims for root bucket freqs")
    ap.add_argument("--time-export", action="store_true")
    ap.add_argument("--time-total", action="store_true")
    ap.add_argument("--stop-expl-mbb", type=float, default=None)
    ap.add_argument("--dense-expl-below-mbb", type=float, default=None)

    args = ap.parse_args()

    t_all0 = time.perf_counter()

    SEED = int(args.seed)
    ITERATIONS = int(args.iters)
    BR_POLICY_ITERS = int(args.br_policy_iters)
    EXP_PRINT_STEP = int(args.exp_print_step)
    EV_SAMPLES_PER_BUCKET = int(args.ev_samples)
    SIM_INIT_FREQ_N = int(args.sim_init_freq_n)

    pots = []
    for tok in str(args.pots).split(","):
        tok = tok.strip()
        if tok:
            pots.append(float(tok))

    random.seed(SEED)
    np.random.seed(SEED)

    print(
        f"1draw CFR+ FULL: iters={ITERATIONS:,} seed={SEED} BB={BIG_BLIND} "
        f"out={args.out_prefix} stop_expl_mbb={args.stop_expl_mbb} "
        f"dense_expl_below_mbb={args.dense_expl_below_mbb}"
    )
    print(f"root freq sims: {SIM_INIT_FREQ_N:,}")

    btn_seeds, btn_w = build_unique_seeds_and_weights_from_grid(btn_grid)
    bb_seeds, bb_w = build_unique_seeds_and_weights_from_grid(bb_grid)

    btn_items = [SeedItem(seed=s, draws=1, weight=float(w)) for s, w in zip(btn_seeds, btn_w)]
    bb_items = [SeedItem(seed=s, draws=1, weight=float(w)) for s, w in zip(bb_seeds, bb_w)]

    print("building bucket-pair chance distribution (exact draw1)...")
    matchups, bucket_freq_by_player = build_bucket_pair_matchups_seedmodel(
        btn_items, bb_items, BUCKETS_1DRAW, bucket_label_1draw, SIM_INIT_FREQ_N
    )
    print(f"done: {len(matchups)} nonzero bucket-pairs")

    pot_list = pots if pots else POT_SIZES

    for pot in pot_list:
        print(f"\npot={pot}")
        random.seed(SEED)
        np.random.seed(SEED)
        reset_infosets()

        tracked_iters, regret_data, evo_data, node_actions, ex_i, ex_mbb, ex_bb, ex_chip = train_cfrplus_sweep(
            matchups=matchups,
            buckets=BUCKETS_1DRAW,
            report_seqs=REPORT_SEQS,
            iterations=ITERATIONS,
            pot=float(pot),
            initial_state_fn=initial_state,
            state_from_history_fn=state_from_history,
            legal_actions_fn=legal_actions,
            step_fn=step,
            ikey_fn=ikey,
            info_key_str_fn=info_key_str,
            br_policy_iters=BR_POLICY_ITERS,
            big_blind=BIG_BLIND,
            exp_print_step=EXP_PRINT_STEP,
            stop_expl_mbb=args.stop_expl_mbb,
            dense_expl_below_mbb=args.dense_expl_below_mbb,
        )

        actual_iterations = ex_i[-1] if ex_i else 0

        export_variant(
            out_prefix=str(args.out_prefix),
            pot=float(pot),
            iterations=actual_iterations,
            buckets=BUCKETS_1DRAW,
            report_seqs=REPORT_SEQS,
            bucket_freq_by_player=bucket_freq_by_player,
            tracked_iters=tracked_iters,
            regret_data=regret_data,
            evo_data=evo_data,
            node_actions=node_actions,
            exploit_iters=ex_i,
            expl_mbb=ex_mbb,
            expl_bb=ex_bb,
            expl_chip=ex_chip,
            initial_state_fn=initial_state,
            state_from_history_fn=state_from_history,
            legal_actions_fn=legal_actions,
            step_fn=step,
            ikey_fn=ikey,
            order_actions_fn=order_actions_limit,
            matchups=matchups,
            ev_samples_per_bucket=EV_SAMPLES_PER_BUCKET,
            time_export=bool(args.time_export),
        )

    print("\n1draw CFR+ done!")
    if args.time_total:
        print(f"[TIME] script_total={time.perf_counter() - t_all0:.2f}s")


if __name__ == "__main__":
    main()
